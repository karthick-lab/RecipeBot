from openpyxl import load_workbook
import os

def get_existing_titles(recipe_path):
    if not os.path.exists(recipe_path):
        return []

    try:
        book = load_workbook(recipe_path)
        return book.sheetnames
    except Exception as e:
        print(f"⚠️ Error reading Excel file: {e}")
        return []