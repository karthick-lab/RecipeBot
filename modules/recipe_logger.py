def log_recipe(raw_text, priced_df, target_path, tray_width, tray_height, num_pieces, serving_size, cost_summary, sheet_name):
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, PatternFill
    import os

    # Load or create workbook
    if os.path.exists(target_path):
        wb = load_workbook(target_path)
    else:
        wb = Workbook()
        default_sheet = wb.active
        if default_sheet.max_row == 1 and default_sheet["A1"].value is None:
            wb.remove(default_sheet)

    # Create new sheet
    ws = wb.create_sheet(title=sheet_name)

    # Styles
    header_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    bold_font = Font(bold=True)

    # Tray and yield info
    ws.append(["Tray Width", tray_width])
    ws.append(["Tray Height", tray_height])
    ws.append(["Yield", num_pieces])
    ws.append(["Serving Size", serving_size])
    for row in ws.iter_rows(min_row=1, max_row=4):
        for cell in row:
            cell.font = bold_font
    ws.append([])

    # Ingredient pricing header
    ws.append(["Ingredient", "Qty", "Unit", "Unit Price (₹)", "Total Cost (₹)"])
    for cell in ws[ws.max_row]:
        cell.fill = header_fill
        cell.font = bold_font

    # Ingredient rows
    start_row = ws.max_row + 1
    for row in priced_df.itertuples(index=False):
        name = row.ingredient
        qty = row.qty
        unit = row.unit

        if qty == 0 and name.lower() == "ingredients":
            continue

        current_row = ws.max_row + 1  # Predict next row before appending

        # Build the price formula
        if unit in ["piece", "pieces"]:
            price_formula = f'=IFERROR(VLOOKUP(A{current_row}, [ingredient_master.xlsx]Sheet1!A:D, 4, FALSE), 0)'
        else:
            price_formula = f'=IFERROR(VLOOKUP(A{current_row}, [ingredient_master.xlsx]Sheet1!A:D, 3, FALSE)/1000, 0)'

        # Append row with formula
        ws.append([name, qty, unit, price_formula, None])
        ws[f"E{current_row}"] = f'=B{current_row} * D{current_row}'

    ws.append([])

    # Cost summary header
    ws.append(["💰 Cost Summary"])
    for cell in ws[ws.max_row]:
        cell.fill = header_fill
        cell.font = bold_font

    summary_start = ws.max_row + 1
    raw_material_formula = f'=SUM(E{start_row}:E{ws.max_row})'
    packing_formula = f'=B3 * 20'  # Yield is in B3

    # Cost breakdown
    ws.append(["Raw Material Cost", raw_material_formula])
    ws.append(["Labor", cost_summary.get("labor", 100)])
    ws.append(["Rent", cost_summary.get("rent", 100)])
    ws.append(["Gas", cost_summary.get("gas", 50)])
    ws.append(["Electricity", cost_summary.get("electricity", 50)])
    ws.append(["Packing", packing_formula])

    total_cost_row = ws.max_row + 1
    ws.append(["Total Manufacturing Cost", f'=SUM(B{summary_start}:B{total_cost_row - 1})'])

    total_cost_per_piece_row = ws.max_row + 1
    ws.append(["Manufacturing Price per piece", f'=B{total_cost_row} / B3'])

    profit_row = ws.max_row + 1
    ws.append(["Profit (25%) per piece or serving", f'=B{total_cost_per_piece_row} * 0.25'])



    selling_price_row = ws.max_row + 1
    ws.append(["Selling Price", f'=B{total_cost_row} + (B{profit_row}*B3)'])

    ws.append(["Selling Price per Piece", f'=B{selling_price_row} / B3'])

    for row in ws.iter_rows(min_row=summary_start - 1, max_row=ws.max_row):
        for cell in row:
            cell.font = bold_font

    ws.append([])

    # Raw recipe section
    ws.append(["📜 Raw Recipe"])
    for cell in ws[ws.max_row]:
        cell.fill = header_fill
        cell.font = bold_font

    for line in raw_text.split("\n"):
        ws.append([line])

    wb.save(target_path)